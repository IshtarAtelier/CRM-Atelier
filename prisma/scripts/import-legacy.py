#!/usr/bin/env python3
"""
import-legacy.py
Importa datos históricos de ATELIER 1.xlsx al CRM.

Modo dry-run (por defecto): solo muestra el matching, no escribe nada.
Modo import: crea Interactions + Tag "Histórico" + actualiza status.

Uso:
  python3 prisma/import-legacy.py                  # dry-run
  python3 prisma/import-legacy.py --execute         # importar de verdad
"""
import sys
import os
import re
import unicodedata
from datetime import datetime
from collections import defaultdict

# ── Dependencias ─────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system("pip3 install openpyxl --quiet")
    import openpyxl

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print("Instalando psycopg2-binary...")
    os.system("pip3 install psycopg2-binary --quiet")
    import psycopg2
    import psycopg2.extras

# ── Config ───────────────────────────────────────────────────────────────────
EXCEL_PATH = os.path.expanduser("~/Downloads/ATELIER 1.xlsx")
DATABASE_URL = "postgresql://postgres:localpassword@localhost:5432/atelier"
DRY_RUN = "--execute" not in sys.argv

# ── Helpers ──────────────────────────────────────────────────────────────────

# Sufijos comunes en nombres del Excel que no están en el CRM
SUFFIX_PATTERNS = [
    r'\s*-\s*(google\s*(ads|maps)?|meta|facebook|instagram|insta|face|calle|whatsapp|wap|cursi\s*(insta|face)?|reel.*|busqueda.*|dr\.?\s*\w+|referid[oa].*|recomendad[oa].*|envio|cliente.*|amig[oa].*|recompra.*|fasta|llamada.*)\s*$',
    r'\s*-\s*(CR|TN)\s*$',
    r'\s*\(.*?\)\s*$',  # Quitar paréntesis como (viejita), (Lili)
    r'\s*//.*$',  # Quitar // y lo que sigue
    r'\s*-\s*$',  # Guión suelto al final
]

def normalize(name: str) -> str:
    """Normaliza un nombre para matching: minúsculas, sin acentos, sin espacios extra."""
    if not name:
        return ""
    # Quitar acentos
    nfkd = unicodedata.normalize('NFKD', name)
    ascii_str = nfkd.encode('ASCII', 'ignore').decode('ASCII')
    # Minúsculas, quitar espacios extra
    result = re.sub(r'\s+', ' ', ascii_str.lower().strip())
    # Quitar emojis y caracteres especiales excepto guiones
    result = re.sub(r'[^\w\s\-/]', '', result)
    return result.strip()

def clean_excel_name(name: str) -> str:
    """Limpia sufijos tipo '- Google Ads', '- Cursi Insta', etc."""
    if not name:
        return ""
    result = name
    for pattern in SUFFIX_PATTERNS:
        result = re.sub(pattern, '', result, flags=re.IGNORECASE)
    return result.strip()

def normalize_phone(phone: str) -> str:
    """Normaliza un teléfono para matching: solo dígitos, últimos 8."""
    if not phone:
        return ""
    digits = re.sub(r'\D', '', phone)
    # Últimos 8 dígitos (sin código de área)
    return digits[-8:] if len(digits) >= 8 else digits

def name_tokens(name: str) -> set:
    """Extrae tokens significativos de un nombre (min 3 chars)."""
    normalized = normalize(name)
    # Quitar sufijos antes de tokenizar
    cleaned = normalize(clean_excel_name(name))
    tokens = set(cleaned.split())
    # Filtrar tokens muy cortos o comunes
    stopwords = {'de', 'del', 'la', 'el', 'los', 'las', 'y', 'e', 'en', 'por', 'con', 'dr', 'casa', 'null'}
    return {t for t in tokens if len(t) >= 3 and t not in stopwords}

def gen_cuid():
    """Genera un ID tipo cuid simple para Prisma."""
    import random
    import string
    import time
    chars = string.ascii_lowercase + string.digits
    timestamp = hex(int(time.time() * 1000))[2:]
    random_part = ''.join(random.choices(chars, k=16))
    return f"c{timestamp}{random_part}"

def format_currency(amount):
    """Formatea un monto como moneda argentina."""
    if not amount or not isinstance(amount, (int, float)):
        return "$0"
    return f"${amount:,.0f}".replace(",", ".")

def format_graduation(eye, esf, cil, eje):
    """Formatea la graduación de un ojo."""
    parts = []
    if esf is not None and esf != 'NULL':
        parts.append(f"ESF {'+' if float(esf) >= 0 else ''}{esf}")
    if cil is not None and cil != 'NULL':
        parts.append(f"CIL {cil}")
    if eje is not None and eje != 'NULL':
        parts.append(f"EJE {eje}°")
    if not parts:
        return ""
    eye_label = f"({eye})" if eye and eye != 'NULL' else ""
    return f" {eye_label}: {' / '.join(parts)}"

# ── Leer Excel ───────────────────────────────────────────────────────────────
print(f"\n📖 Leyendo {EXCEL_PATH}...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# --- Clientes ---
ws_clientes = wb['CLIENTES']
excel_clients = []
for row in ws_clientes.iter_rows(min_row=2, max_row=ws_clientes.max_row, values_only=True):
    nombre = row[0]
    if not nombre:
        continue
    excel_clients.append({
        'nombre': nombre,
        'telefono': str(row[6]) if row[6] and row[6] != 'NULL' else None,
        'email': row[10] if row[10] and row[10] != 'NULL' else None,
        'direccion': row[5] if row[5] and row[5] != 'NULL' else None,
        'cobertura': row[11] if row[11] and row[11] != 'NULL' and row[11] != 0.0 else None,
        'observaciones': row[13] if row[13] and row[13] != 'NULL' else None,
        'created_at': row[14],
    })

# --- Ventas ---
ws_ventas = wb['VENTAS']
ventas = {}
for row in ws_ventas.iter_rows(min_row=2, max_row=ws_ventas.max_row, values_only=True):
    num_venta = int(row[0]) if row[0] else None
    if num_venta is None:
        continue
    ventas[num_venta] = {
        'num': num_venta,
        'nombre': row[1],
        'fecha': row[2],
        'total': row[3] if isinstance(row[3], (int, float)) else 0,
        'items': [],
    }

# --- Detalle Ventas ---
ws_detalle = wb['DETALLE VENTAS']
for row in ws_detalle.iter_rows(min_row=2, max_row=ws_detalle.max_row, values_only=True):
    num_venta = int(row[0]) if row[0] else None
    if num_venta is None or num_venta not in ventas:
        continue
    ventas[num_venta]['items'].append({
        'ojo': row[1] if row[1] and row[1] != 'NULL' else None,
        'articulo': row[2] if row[2] and row[2] != 'NULL' else 'Artículo sin nombre',
        'esf': row[3] if row[3] and row[3] != 'NULL' else None,
        'cil': row[4] if row[4] and row[4] != 'NULL' else None,
        'eje': row[5] if row[5] and row[5] != 'NULL' else None,
        'cantidad': int(row[6]) if row[6] else 1,
        'precio': row[7] if isinstance(row[7], (int, float)) else 0,
    })

# Agrupar ventas por nombre de cliente
ventas_por_cliente = defaultdict(list)
for v in ventas.values():
    if v['nombre']:
        ventas_por_cliente[v['nombre']].append(v)

print(f"   ✅ {len(excel_clients)} clientes, {len(ventas)} ventas, {sum(len(v['items']) for v in ventas.values())} líneas de detalle")

# ── Conectar a PostgreSQL ────────────────────────────────────────────────────
print(f"\n🔗 Conectando a la base de datos...")
conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# Obtener todos los clientes del CRM
cur.execute('SELECT id, name, phone, email, status, address, insurance, "contactSource" FROM "Client"')
crm_clients = cur.fetchall()
print(f"   ✅ {len(crm_clients)} clientes en el CRM actual")

# ── Matching (multi-estrategia) ──────────────────────────────────────────────
print(f"\n🔍 Matcheando clientes (3 estrategias)...")

# Índice 1: Nombre normalizado exacto
crm_by_name = {}
for c in crm_clients:
    key = normalize(c['name'])
    if key:
        crm_by_name[key] = c

# Índice 2: Teléfono (últimos 8 dígitos)
crm_by_phone = {}
for c in crm_clients:
    if c.get('phone'):
        phone_key = normalize_phone(c['phone'])
        if phone_key and len(phone_key) >= 7:
            crm_by_phone[phone_key] = c

# Índice 3: Tokens del nombre (para fuzzy matching)
crm_by_tokens = {}
for c in crm_clients:
    tokens = name_tokens(c['name'])
    if len(tokens) >= 2:
        frozen = frozenset(tokens)
        crm_by_tokens[frozen] = c

def find_match(ec):
    """Intenta matchear un cliente del Excel con el CRM usando 3 estrategias."""
    # Estrategia 1: Nombre exacto normalizado
    key = normalize(ec['nombre'])
    if key in crm_by_name:
        return crm_by_name[key], "nombre_exacto"
    
    # Estrategia 2: Nombre limpio (sin sufijos) normalizado
    cleaned = normalize(clean_excel_name(ec['nombre']))
    if cleaned and cleaned in crm_by_name:
        return crm_by_name[cleaned], "nombre_limpio"
    
    # Estrategia 3: Teléfono
    if ec.get('telefono'):
        phone_key = normalize_phone(ec['telefono'])
        if phone_key and len(phone_key) >= 7 and phone_key in crm_by_phone:
            return crm_by_phone[phone_key], "telefono"
    
    # Estrategia 4: Tokens del nombre (todas las palabras del Excel están en el CRM)
    excel_tokens = name_tokens(ec['nombre'])
    if len(excel_tokens) >= 2:
        for crm_tokens, crm_client in crm_by_tokens.items():
            # El nombre del CRM tiene que contener todos los tokens significativos del Excel (limpio)
            clean_tokens = name_tokens(clean_excel_name(ec['nombre']))
            if clean_tokens and len(clean_tokens) >= 2 and clean_tokens == crm_tokens:
                return crm_client, "tokens"
    
    return None, None

# Matchear
matched = []
unmatched = []
unmatched_with_sales = []
match_methods = defaultdict(int)
already_matched_crm_ids = set()  # Evitar duplicados

for ec in excel_clients:
    ventas_cliente = ventas_por_cliente.get(ec['nombre'], [])
    total_ventas = sum(v['total'] for v in ventas_cliente)
    
    crm_match, method = find_match(ec)
    
    if crm_match and crm_match['id'] not in already_matched_crm_ids:
        already_matched_crm_ids.add(crm_match['id'])
        match_methods[method] += 1
        matched.append({
            'excel': ec,
            'crm': crm_match,
            'ventas': ventas_cliente,
            'total': total_ventas,
            'method': method,
        })
    else:
        if ventas_cliente:
            unmatched_with_sales.append({
                'excel': ec,
                'ventas': ventas_cliente,
                'total': total_ventas,
            })
        else:
            unmatched.append(ec)

print(f"\n{'='*60}")
print(f"📊 RESULTADO DEL MATCHING")
print(f"{'='*60}")
print(f"   ✅ Matcheados:              {len(matched)}")
print(f"   ❌ Sin match (con ventas):   {len(unmatched_with_sales)}")
print(f"   ⚪ Sin match (sin ventas):   {len(unmatched)}")
print(f"{'='*60}")

# Matcheados con ventas
matched_with_sales = [m for m in matched if m['ventas']]
print(f"\n   De los matcheados, {len(matched_with_sales)} tienen historial de ventas")
total_historico = sum(m['total'] for m in matched_with_sales)
print(f"   Total histórico de los matcheados: {format_currency(total_historico)}")

# Mostrar los no matcheados con ventas (importante!)
if unmatched_with_sales:
    print(f"\n⚠️  CLIENTES CON VENTAS QUE NO MATCHEARON (revisión manual):")
    for u in sorted(unmatched_with_sales, key=lambda x: -x['total']):
        print(f"   • {u['excel']['nombre']} — {format_currency(u['total'])} ({len(u['ventas'])} ventas)")


# ── Funciones de formato ─────────────────────────────────────────────────────

def build_note(match_data):
    """Construye la nota de historial para un cliente."""
    ventas_list = sorted(match_data['ventas'], key=lambda v: v['fecha'] if v['fecha'] else datetime.min)
    total = match_data['total']
    
    lines = [
        "📋 HISTORIAL SISTEMA ANTERIOR (Jul 2023 - Mar 2025)",
        "═" * 50,
        "",
        f"💰 Total gastado: {format_currency(total)} ({len(ventas_list)} {'compra' if len(ventas_list) == 1 else 'compras'})",
        "",
    ]
    
    for v in ventas_list:
        fecha_str = v['fecha'].strftime('%d/%m/%Y') if v['fecha'] else '??/??/????'
        lines.append(f"📦 Venta #{v['num']} — {fecha_str} — {format_currency(v['total'])}")
        
        for item in v['items']:
            grad = format_graduation(item['ojo'], item['esf'], item['cil'], item['eje'])
            qty = f"×{item['cantidad']}" if item['cantidad'] > 1 else ""
            lines.append(f"   • {item['articulo']} {qty} — {format_currency(item['precio'])}{grad}")
        
        lines.append("")
    
    # Agregar observaciones del sistema anterior si las tiene
    obs = match_data['excel'].get('observaciones')
    if obs:
        lines.extend([
            "📝 Observaciones del sistema anterior:",
            f"   {obs}",
            "",
        ])
    
    return "\n".join(lines)

# ── DRY-RUN o IMPORTAR ──────────────────────────────────────────────────────

if DRY_RUN:
    print(f"\n🔒 MODO DRY-RUN — No se escribió nada en la base de datos.")
    print(f"   Para importar, ejecutá: python3 prisma/import-legacy.py --execute")
    
    # Mostrar preview de las primeras 3 notas que se crearían
    print(f"\n📝 PREVIEW — Así se verían las notas para los primeros 3 clientes con ventas:")
    for m in matched_with_sales[:3]:
        note = build_note(m)
        print(f"\n{'─'*50}")
        print(f"Cliente: {m['crm']['name']} (CRM ID: {m['crm']['id']})")
        print(f"{'─'*50}")
        print(note[:500])
        if len(note) > 500:
            print(f"... ({len(note)} caracteres en total)")
    
    conn.close()
    sys.exit(0)

# ── IMPORTAR ─────────────────────────────────────────────────────────────────
print(f"\n🚀 IMPORTANDO...")

# 1. Crear o buscar tag "Histórico"
cur.execute('SELECT id FROM "Tag" WHERE name = %s', ('Histórico',))
tag_row = cur.fetchone()
if tag_row:
    tag_id = tag_row['id']
    print(f"   Tag 'Histórico' ya existe: {tag_id}")
else:
    tag_id = gen_cuid()
    cur.execute(
        'INSERT INTO "Tag" (id, name, color, "botAction") VALUES (%s, %s, %s, %s)',
        (tag_id, 'Histórico', '#8B7355', 'NONE')
    )
    print(f"   ✅ Tag 'Histórico' creado: {tag_id}")

# 2. Crear Interactions y actualizar clientes
interactions_created = 0
clients_updated = 0
tags_assigned = 0

for m in matched_with_sales:
    client_id = m['crm']['id']
    
    # Crear la nota de historial
    note_content = build_note(m)
    interaction_id = gen_cuid()
    
    cur.execute(
        'INSERT INTO "Interaction" (id, "clientId", type, content, "createdAt") VALUES (%s, %s, %s, %s, %s)',
        (interaction_id, client_id, 'NOTE', note_content, datetime.now())
    )
    interactions_created += 1
    
    # Actualizar status a CLIENT si tiene ventas y no es ya CLIENT
    if m['crm']['status'] != 'CLIENT':
        cur.execute(
            'UPDATE "Client" SET status = %s, "contactSource" = COALESCE("contactSource", %s), "updatedAt" = %s WHERE id = %s',
            ('CLIENT', 'Sistema Anterior', datetime.now(), client_id)
        )
        clients_updated += 1
    
    # Asignar tag "Histórico" (tabla de relación many-to-many)
    # Prisma usa una tabla implícita _ClientToTag
    cur.execute(
        'SELECT 1 FROM "_ClientToTag" WHERE "A" = %s AND "B" = %s',
        (client_id, tag_id)
    )
    if not cur.fetchone():
        cur.execute(
            'INSERT INTO "_ClientToTag" ("A", "B") VALUES (%s, %s)',
            (client_id, tag_id)
        )
        tags_assigned += 1

# 3. Para clientes matcheados SIN ventas pero con observaciones, agregar la nota
obs_only = 0
for m in matched:
    if m['ventas']:
        continue  # ya procesado
    obs = m['excel'].get('observaciones')
    if obs:
        client_id = m['crm']['id']
        interaction_id = gen_cuid()
        note = f"📝 Observaciones del sistema anterior:\n{obs}"
        cur.execute(
            'INSERT INTO "Interaction" (id, "clientId", type, content, "createdAt") VALUES (%s, %s, %s, %s, %s)',
            (interaction_id, client_id, 'NOTE', note, datetime.now())
        )
        obs_only += 1

conn.commit()

print(f"\n{'='*60}")
print(f"✅ IMPORTACIÓN COMPLETADA")
print(f"{'='*60}")
print(f"   📝 Interactions creadas (historial):  {interactions_created}")
print(f"   📝 Interactions creadas (solo obs):    {obs_only}")
print(f"   👤 Clientes actualizados a CLIENT:     {clients_updated}")
print(f"   🏷️  Tags 'Histórico' asignados:        {tags_assigned}")
print(f"{'='*60}")

conn.close()
