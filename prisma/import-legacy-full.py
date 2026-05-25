#!/usr/bin/env python3
"""
import-legacy-full.py
Cruza ATELIER 1.xlsx y ATELIER 2.xlsx.
Limpia datos anteriores.
Crea clientes faltantes y actualiza existentes.
Genera historiales unificados.
"""
import sys
import os
import re
import unicodedata
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    os.system("pip3 install openpyxl --quiet")
    import openpyxl

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    os.system("pip3 install psycopg2-binary --quiet")
    import psycopg2
    import psycopg2.extras

DATABASE_URL = os.environ.get('DATABASE_URL', "postgresql://postgres:localpassword@localhost:5432/atelier")
DRY_RUN = "--execute" not in sys.argv

# ── Helpers ──
def normalize(name):
    if not name: return ''
    nfkd = unicodedata.normalize('NFKD', str(name))
    ascii_str = nfkd.encode('ASCII', 'ignore').decode('ASCII')
    result = re.sub(r'\s+', ' ', ascii_str.lower().strip())
    result = re.sub(r'[^\w\s\-/]', '', result)
    return result.strip()

SUFFIX_PATTERNS = [
    r'\s*-\s*(google\s*(ads|maps)?|meta|facebook|instagram|insta|face|calle|whatsapp|wap|cursi\s*(insta|face)?|reel.*|busqueda.*|dr\.?\s*\w+|referid[oa].*|recomendad[oa].*|envio|cliente.*|amig[oa].*|recompra.*|fasta|llamada.*)\s*$',
    r'\s*-\s*(CR|TN)\s*$',
    r'\s*//.*$',
    r'\s*-\s*$',
]
PAREN_PATTERN = r'\s*\(.*?\)\s*$'

def extract_source(name):
    if not name: return None
    name = str(name)
    for p in SUFFIX_PATTERNS:
        m = re.search(p, name, re.IGNORECASE)
        if m:
            return m.group(0).strip().lstrip('- ').strip()
    return None

def clean_name(name):
    if not name: return ''
    result = str(name)
    for p in SUFFIX_PATTERNS:
        result = re.sub(p, '', result, flags=re.IGNORECASE)
    result = re.sub(PAREN_PATTERN, '', result)
    return result.strip()

def normalize_phone(phone):
    if not phone: return ''
    digits = re.sub(r'\D', '', str(phone))
    return digits[-8:] if len(digits) >= 8 else digits

def name_tokens(name):
    cleaned = normalize(clean_name(name))
    tokens = set(cleaned.split())
    stopwords = {'de', 'del', 'la', 'el', 'los', 'las', 'y', 'e', 'en', 'por', 'con', 'dr', 'casa', 'null', '-'}
    return {t for t in tokens if len(t) >= 3 and t not in stopwords}

def gen_cuid():
    import random, string, time
    chars = string.ascii_lowercase + string.digits
    timestamp = hex(int(time.time() * 1000))[2:]
    random_part = ''.join(random.choices(chars, k=16))
    return f"c{timestamp}{random_part}"

def format_currency(amount):
    if not amount or not isinstance(amount, (int, float)): return "$0"
    return f"${amount:,.0f}".replace(",", ".")

# ── Leer Ambos Excels ──
clientes_excel = defaultdict(dict)
ventas_excel = defaultdict(list)

def process_excel(filepath, prefix):
    print(f'📖 Procesando {filepath}...')
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Clientes
    ws_c = wb['CLIENTES']
    col_tel = 6 if prefix == 'AT1' else 12
    col_email = 10 if prefix == 'AT1' else 13
    col_obs = 13 if prefix == 'AT1' else 17
    
    for row in ws_c.iter_rows(min_row=2, max_row=ws_c.max_row, values_only=True):
        nombre = row[0]
        if not nombre: continue
        key = normalize(clean_name(nombre))
        if key not in clientes_excel:
            clientes_excel[key] = {
                'nombre_original': str(nombre),
                'nombre_limpio': clean_name(nombre).title(),
                'origen': extract_source(nombre),
                'telefono': str(row[col_tel]) if row[col_tel] and str(row[col_tel]) not in ['NULL', '\\N', 'None'] else None,
                'email': str(row[col_email]) if row[col_email] and str(row[col_email]) not in ['NULL', '\\N', 'None'] else None,
                'observaciones': [str(row[col_obs])] if row[col_obs] and str(row[col_obs]) not in ['NULL', '\\N', 'None'] else [],
                'fuentes': [prefix]
            }
        else:
            if row[col_tel] and str(row[col_tel]) not in ['NULL', '\\N', 'None']: clientes_excel[key]['telefono'] = str(row[col_tel])
            if row[col_email] and str(row[col_email]) not in ['NULL', '\\N', 'None']: clientes_excel[key]['email'] = str(row[col_email])
            if row[col_obs] and str(row[col_obs]) not in ['NULL', '\\N', 'None'] and str(row[col_obs]) not in clientes_excel[key]['observaciones']:
                clientes_excel[key]['observaciones'].append(str(row[col_obs]))
            if prefix not in clientes_excel[key]['fuentes']: clientes_excel[key]['fuentes'].append(prefix)

    # Ventas
    ws_v = wb['VENTAS']
    col_v_num = 0
    col_v_nombre = 1 if prefix == 'AT1' else 2
    col_v_fecha = 2 if prefix == 'AT1' else 1
    col_v_total = 3 if prefix == 'AT1' else 7
    
    ventas_dict = {}
    for row in ws_v.iter_rows(min_row=2, max_row=ws_v.max_row, values_only=True):
        nv = row[col_v_num]
        if not nv: continue
        nombre = row[col_v_nombre]
        if not nombre: continue
        key = normalize(clean_name(nombre))
        v_id = f'{prefix}-{nv}'
        ventas_dict[v_id] = {
            'num': v_id,
            'nombre_original': str(nombre),
            'cliente_key': key,
            'fecha': row[col_v_fecha] if isinstance(row[col_v_fecha], datetime) else None,
            'total': row[col_v_total] if isinstance(row[col_v_total], (int, float)) else 0,
            'items': [],
            'origen_archivo': prefix
        }

    # Detalle Ventas
    ws_d = wb['DETALLE VENTAS']
    col_d_venta = 0
    col_d_ojo = 1 if prefix == 'AT1' else 2
    col_d_art = 2 if prefix == 'AT1' else 3
    col_d_esf = 3 if prefix == 'AT1' else 4
    col_d_cil = 4 if prefix == 'AT1' else 5
    col_d_eje = 5 if prefix == 'AT1' else 6
    col_d_cant = 6 if prefix == 'AT1' else 10
    col_d_precio = 7 if prefix == 'AT1' else 11

    for row in ws_d.iter_rows(min_row=2, max_row=ws_d.max_row, values_only=True):
        nv = row[col_d_venta]
        if not nv: continue
        v_id = f'{prefix}-{nv}'
        if v_id not in ventas_dict: continue
        
        ventas_dict[v_id]['items'].append({
            'ojo': row[col_d_ojo] if row[col_d_ojo] and row[col_d_ojo] != 'NULL' else None,
            'articulo': str(row[col_d_art]) if row[col_d_art] and row[col_d_art] != 'NULL' else '?',
            'esf': row[col_d_esf] if row[col_d_esf] and row[col_d_esf] != 'NULL' else None,
            'cil': row[col_d_cil] if row[col_d_cil] and row[col_d_cil] != 'NULL' else None,
            'eje': row[col_d_eje] if row[col_d_eje] and row[col_d_eje] != 'NULL' else None,
            'cantidad': int(row[col_d_cant]) if row[col_d_cant] else 1,
            'precio': row[col_d_precio] if isinstance(row[col_d_precio], (int, float)) else 0,
        })

    for v in ventas_dict.values():
        ventas_excel[v['cliente_key']].append(v)

base_dir = os.path.dirname(os.path.abspath(__file__))
path_at1 = os.path.join(base_dir, 'legacy_data', 'ATELIER 1.xlsx')
path_at2 = os.path.join(base_dir, 'legacy_data', 'ATELIER 2.xlsx')

process_excel(path_at1, 'AT1')
process_excel(path_at2, 'AT2')

# ── CRM Matching ──
print(f"\n🔗 Conectando a la base de datos...")
conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur.execute('SELECT id, name, phone, email, status, "contactSource" FROM "Client"')
crm_clients = cur.fetchall()

crm_by_name = {}
for c in crm_clients:
    if key := normalize(c['name']): crm_by_name[key] = c

crm_by_phone = {}
for c in crm_clients:
    if c.get('phone'):
        if pk := normalize_phone(c['phone']):
            if len(pk) >= 7: crm_by_phone[pk] = c

crm_by_tokens = {}
for c in crm_clients:
    tokens = name_tokens(c['name'])
    if len(tokens) >= 2: crm_by_tokens[frozenset(tokens)] = c

def find_match(client_key, ec):
    if client_key in crm_by_name: return crm_by_name[client_key], 'nombre_limpio'
    key_orig = normalize(ec['nombre_original'])
    if key_orig in crm_by_name: return crm_by_name[key_orig], 'nombre_exacto'
    
    if ec.get('telefono'):
        if pk := normalize_phone(ec['telefono']):
            if len(pk) >= 7 and pk in crm_by_phone: return crm_by_phone[pk], 'telefono'
            
    clean_tokens = name_tokens(ec['nombre_limpio'])
    if len(clean_tokens) >= 2:
        for ct, cc in crm_by_tokens.items():
            if clean_tokens == ct: return cc, 'tokens'
    return None, None

matched = []
unmatched = []

for key, ec in clientes_excel.items():
    vs = ventas_excel.get(key, [])
    # Filtro: descartar clientes sin info valiosa
    if not vs and not ec['observaciones']:
        continue
        
    crm, method = find_match(key, ec)
    if crm:
        matched.append({'key': key, 'excel': ec, 'ventas': vs, 'crm': crm})
    else:
        unmatched.append({'key': key, 'excel': ec, 'ventas': vs})

def format_graduation(eye, esf, cil, eje):
    parts = []
    if esf is not None and str(esf) not in ['NULL', '\\N', 'None']:
        try:
            val = float(esf)
            parts.append(f"ESF {'+' if val >= 0 else ''}{esf}")
        except ValueError:
            parts.append(f"ESF {esf}")
    if cil is not None and str(cil) not in ['NULL', '\\N', 'None']: parts.append(f"CIL {cil}")
    if eje is not None and str(eje) not in ['NULL', '\\N', 'None']: parts.append(f"EJE {eje}°")
    if not parts: return ""
    eye_label = f"({eye})" if eye and eye != 'NULL' else ""
    return f" {eye_label}: {' / '.join(parts)}"

def build_note(ec, vs):
    vs_sorted = sorted(vs, key=lambda v: v['fecha'] if v['fecha'] else datetime.min)
    total = sum(v['total'] for v in vs)
    
    lines = [
        "📋 HISTORIAL SISTEMA ANTERIOR (ATELIER 1 + 2)",
        "═" * 50,
        "",
        f"💰 Total gastado: {format_currency(total)} ({len(vs_sorted)} {'compra' if len(vs_sorted) == 1 else 'compras'})",
        "",
    ]
    
    for v in vs_sorted:
        fecha_str = v['fecha'].strftime('%d/%m/%Y') if v['fecha'] else '??/??/????'
        lines.append(f"📦 Venta #{v['num']} — {fecha_str} — {format_currency(v['total'])}")
        
        for item in v['items']:
            grad = format_graduation(item['ojo'], item['esf'], item['cil'], item['eje'])
            qty = f"×{item['cantidad']}" if item['cantidad'] > 1 else ""
            lines.append(f"   • {item['articulo']} {qty} — {format_currency(item['precio'])}{grad}")
        lines.append("")
        
    if ec['observaciones']:
        lines.append("📝 Observaciones del sistema anterior:")
        for obs in ec['observaciones']:
            lines.append(f"   - {obs}")
        lines.append("")
        
    return "\n".join(lines)


if DRY_RUN:
    print(f"\n🔒 MODO DRY-RUN — No se escribió nada.")
    print(f"Para importar: python3 prisma/import-legacy-full.py --execute")
    conn.close()
    sys.exit(0)

# ── EJECUCIÓN ──
print(f"\n🚀 EJECUTANDO IMPORTACIÓN MASIVA...")

# 1. Limpiar importaciones previas de hoy
cur.execute("DELETE FROM \"Interaction\" WHERE content LIKE '%HISTORIAL SISTEMA ANTERIOR%'")
print(f"   🧹 Borradas las interactions viejas de prueba: {cur.rowcount}")

# 2. Obtener Tag Histórico
cur.execute('SELECT id FROM "Tag" WHERE name = %s', ('Histórico',))
tag_row = cur.fetchone()
if tag_row:
    tag_id = tag_row['id']
else:
    tag_id = gen_cuid()
    cur.execute('INSERT INTO "Tag" (id, name, color, "botAction") VALUES (%s, %s, %s, %s)',
                (tag_id, 'Histórico', '#8B7355', 'NONE'))

stats = {'creados': 0, 'actualizados': 0, 'notas': 0, 'tags': 0}

# 3. Crear Nuevos Clientes
new_clients_map = {}
for u in unmatched:
    ec = u['excel']
    vs = u['ventas']
    cuid = gen_cuid()
    
    status = 'CLIENT' if vs else 'CONTACT'
    origen = ec['origen'] if ec['origen'] else 'Sistema Anterior'
    
    cur.execute('''
        INSERT INTO "Client" (id, name, phone, email, status, "contactSource", "createdAt", "updatedAt")
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    ''', (
        cuid,
        ec['nombre_limpio'] or ec['nombre_original'],
        ec['telefono'],
        None,  # No importamos email para evitar conflictos de constraint Unique
        status,
        origen,
        datetime.now(),
        datetime.now()
    ))
    stats['creados'] += 1
    new_clients_map[u['key']] = cuid

# 4. Actualizar Clientes Existentes
for m in matched:
    crm = m['crm']
    ec = m['excel']
    vs = m['ventas']
    client_id = crm['id']
    
    new_status = 'CLIENT' if vs and crm['status'] != 'CLIENT' else crm['status']
    new_source = crm['contactSource']
    if not new_source and ec['origen']:
        new_source = ec['origen']
        
    if new_status != crm['status'] or new_source != crm['contactSource']:
        cur.execute('''
            UPDATE "Client" 
            SET status = %s, "contactSource" = %s, "updatedAt" = %s
            WHERE id = %s
        ''', (new_status, new_source, datetime.now(), client_id))
        stats['actualizados'] += 1

# 5. Generar Notas y Tags para todos
all_to_process = []
for m in matched: all_to_process.append((m['crm']['id'], m['excel'], m['ventas']))
for u in unmatched: all_to_process.append((new_clients_map[u['key']], u['excel'], u['ventas']))

for client_id, ec, vs in all_to_process:
    # Nota
    note_content = build_note(ec, vs)
    inter_id = gen_cuid()
    cur.execute('''
        INSERT INTO "Interaction" (id, "clientId", type, content, "createdAt")
        VALUES (%s, %s, %s, %s, %s)
    ''', (inter_id, client_id, 'NOTE', note_content, datetime.now()))
    stats['notas'] += 1
    
    # Tag
    cur.execute('SELECT 1 FROM "_ClientToTag" WHERE "A" = %s AND "B" = %s', (client_id, tag_id))
    if not cur.fetchone():
        cur.execute('INSERT INTO "_ClientToTag" ("A", "B") VALUES (%s, %s)', (client_id, tag_id))
        stats['tags'] += 1

conn.commit()
print(f"\n✅ IMPORTACIÓN EXITOSA")
print(f"   👤 Nuevos clientes creados:    {stats['creados']}")
print(f"   🔄 Clientes actualizados:      {stats['actualizados']}")
print(f"   📝 Notas de historial creadas: {stats['notas']}")
print(f"   🏷️  Tags asignados:             {stats['tags']}")
conn.close()
